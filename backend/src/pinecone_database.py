import os
from pinecone import Pinecone
from docx import Document
from openpyxl import load_workbook

pc = Pinecone(api_key="xxxx")
index_name = "ai-receptionist"

if index_name in pc.list_indexes():
    print(f"Index '{index_name}' already exists. Deleting it to recreate...")
    pc.delete_index(index_name)
    print(f"Index '{index_name}' deleted successfully.")

print(f"Creating index '{index_name}'...")
pc.create_index(index_name, dimension=1536)
print(f"Index '{index_name}' created successfully.")

def extract_text_from_xlsx(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    text = []
    for row in sheet.iter_rows(values_only=True):
        text.append(" ".join([str(cell) for cell in row if cell is not None]))
    return "\n".join(text)

def extract_text_from_docx(file_path):
    doc = Document(file_path)
    text = []
    for para in doc.paragraphs:
        text.append(para.text)
    return "\n".join(text)
working_dir = "."

for filename in os.listdir(working_dir):
    file_path = os.path.join(working_dir, filename)
    
    if filename.endswith('.xlsx'):
        print(f"Processing {filename} (xlsx)")
        file_text = extract_text_from_xlsx(file_path)
    
    elif filename.endswith('.docx'):
        print(f"Processing {filename} (docx)")
        file_text = extract_text_from_docx(file_path)
    
    else:
        continue
    
    embeddings = pc.inference.embed(
        model="multilingual-e5-large",
        inputs=[file_text],
        parameters={"input_type": "passage", "truncate": "END"}
    )
    
    index = pc.Index(index_name)
    for i, embedding in enumerate(embeddings):
        print("embeddings:", embeddings)
        index.upsert(
            vectors=[{
                'id': f"file_{filename}_vec_{i}",
                'values': embedding['values'],
                'metadata': {'source': filename, 'text': file_text}
            }]
        )

    print(f"Finished processing {filename}")